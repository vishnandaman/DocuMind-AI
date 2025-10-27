import os
import uuid
from pathlib import Path
from typing import Dict, List, Any
import PyPDF2
from docx import Document as DocxDocument
import openpyxl
import pandas as pd
from datetime import datetime
import re

class DocumentProcessor:
    def __init__(self):
        self.supported_formats = {
            '.pdf': self._process_pdf,
            '.docx': self._process_docx,
            '.txt': self._process_txt,
            '.xlsx': self._process_xlsx,
            '.csv': self._process_csv
        }
    
    async def process_document(self, file_path: Path, filename: str) -> Dict[str, Any]:
        """Process a document and extract text content"""
        file_extension = Path(filename).suffix.lower()
        
        if file_extension not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {file_extension}")
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        # Extract text content
        content = await self.supported_formats[file_extension](file_path)
        
        # Create metadata
        metadata = {
            'filename': filename,
            'file_type': file_extension,
            'file_size': file_size,
            'upload_date': datetime.utcnow().isoformat(),
            'author': None,
            'title': None,
            'summary': None
        }
        
        # Split content into chunks
        chunks = self._split_into_chunks(content)
        
        return {
            'content': content,
            'metadata': metadata,
            'chunks': chunks
        }
    
    def _split_into_chunks(self, content: str, chunk_size: int = 500, overlap: int = 100) -> List[str]:
        """Split content into overlapping chunks for better search"""
        if len(content) <= chunk_size:
            return [content]
        
        chunks = []
        start = 0
        
        while start < len(content):
            end = start + chunk_size
            
            # Try to break at sentence boundary
            if end < len(content):
                # Look for sentence endings
                sentence_end = content.rfind('.', start, end)
                if sentence_end > start + chunk_size // 2:
                    end = sentence_end + 1
                else:
                    # Look for line breaks
                    line_end = content.rfind('\n', start, end)
                    if line_end > start + chunk_size // 2:
                        end = line_end + 1
            
            chunk = content[start:end].strip()
            if chunk and len(chunk) > 50:  # Only add chunks with meaningful content
                chunks.append(chunk)
            
            start = end - overlap
            if start >= len(content):
                break
        
        print(f"Created {len(chunks)} chunks from content of length {len(content)}")
        return chunks
    
    async def _process_pdf(self, file_path: Path) -> str:
        """Extract text from PDF"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                
                return text.strip()
        except Exception as e:
            raise ValueError(f"Error processing PDF: {str(e)}")
    
    async def _process_docx(self, file_path: Path) -> str:
        """Extract text from DOCX"""
        try:
            doc = DocxDocument(file_path)
            text = ""
            
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            
            return text.strip()
        except Exception as e:
            raise ValueError(f"Error processing DOCX: {str(e)}")
    
    async def _process_txt(self, file_path: Path) -> str:
        """Extract text from TXT"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read().strip()
        except Exception as e:
            raise ValueError(f"Error processing TXT: {str(e)}")
    
    async def _process_xlsx(self, file_path: Path) -> str:
        """Extract text from XLSX"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                
                text += "\n"
            
            return text.strip()
        except Exception as e:
            raise ValueError(f"Error processing XLSX: {str(e)}")
    
    async def _process_csv(self, file_path: Path) -> str:
        """Extract text from CSV"""
        try:
            df = pd.read_csv(file_path)
            text = ""
            
            # Add column headers
            text += " | ".join(df.columns) + "\n"
            
            # Add data rows
            for _, row in df.iterrows():
                row_text = " | ".join([str(cell) if pd.notna(cell) else "" for cell in row])
                text += row_text + "\n"
            
            return text.strip()
        except Exception as e:
            raise ValueError(f"Error processing CSV: {str(e)}")
